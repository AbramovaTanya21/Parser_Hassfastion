from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import os
import openpyxl
from openpyxl import Workbook,load_workbook
import threading
import queue
import requests
import json
from PIL import Image
from io import BytesIO






# Указание пути компонентам парсера
# Определение пути к chromedriver 
ChromedriverPuth = 'G:\\Applications\\Selenium\\Chrome_draver\\chromedriver-win64\\142\\chromedriver.exe'
# Определение пути к файлу с коллекциями HassCatalog   
CollectionFilePath = "G:\\Program Proects\\Parsers\\Parser_HassFashion\\Parser_HassFashoin\\Parser_HassFashoin\\HassCatalog.xlsx" 
# Определение имени и автоформирование относительного пути к файлу загрузки данных HassDate
file_name = "HassDate.xlsx"
file_path = "./" + file_name  

# Авторизация на ХассФэсшен 
def Authorization():      
    a=Service(ChromedriverPuth) # Создание экземпляра драйвера
    driver = webdriver.Chrome(service=a)
    driver.get("https://hassfashion.ru/auth/?login=yes")    
    Auth_Log = driver.find_element(By.NAME, "USER_LOGIN")
    Auth_Log.send_keys('Olga_Guschina@mail.ru')
    Auth_Password = driver.find_element(By.NAME, "USER_PASSWORD")
    Auth_Password.send_keys('optopt')
    Auth_Imp = driver.find_element(By.NAME, "Login")
    Auth_Imp.click() 
    time.sleep(5) 
    # with open('cookies.json', 'w') as f: # Создание куки
    #     json.dump(s.cookies.get_dict(), f)

#Получение страниц каталога из Экселя
def GettingLinks(LinksLoaded):   
    NotCount = 0
    #Создание экземпляра драйвера
    s=Service(ChromedriverPuth) 
    driver = webdriver.Chrome(service=s)
    # Обращение к эксель файлу 
    wb = load_workbook(CollectionFilePath)
    ws = wb.active 
    LinkPages = [] 
    last_collection = ws[2][0].value  
    for row in range(2, ws.max_row+1):  
        current_collection = ws[row][0].value 
        if current_collection == None: break  
        if last_collection !=  current_collection : 
             ParsingCollection(driver, LinksLoaded, LinkPages, last_collection, NotCount)             
        LinkPages.append(ws[row][2].value) 
        last_collection = current_collection
    ParsingCollection(driver, LinksLoaded, LinkPages, last_collection, NotCount) 
    driver.quit()

def ParsingCollection(driver, LinksLoaded, LinkPages, last_collection, NotCount):
   print(f'GettingLinks: Начат сбор ссылок для коллекции {last_collection}')  
   Links = [] 
   FilteredPaginatorConteyner = []
   for LinkPage in LinkPages: 
        driver.get(LinkPage)  
        PaginatorPage = LinkPage 
        while True:
            # Отбор товаров в наличие                        
            OnSales = driver.find_elements(By.XPATH, "//div[@class='res_cards']/div[not(@style)]") 
            for OnSale in OnSales:   
                a = OnSale.find_element(By.TAG_NAME, "a")
                Links.append(a.get_attribute("href"))     
            try:
                if LinkPage.find("?"):
                    PaginatorLoadPage = LinkPage + "&" + driver.find_element(By.XPATH, "//div[contains(@data-pagination,'PAGEN')]").get_attribute("data-pagination") 
                else:
                    PaginatorLoadPage = LinkPage + "?" + driver.find_element(By.XPATH, "//div[contains(@data-pagination,'PAGEN')]").get_attribute("data-pagination")
                driver.get(PaginatorLoadPage)           
            except: break        
   LinkPages.clear()        
   print(f'GettingLinks: Закончен сбор ссылок для коллекции {last_collection}')
   # GoodsLinksQueue.put((last_collection, Links))
   GoodsLinksQueue.put((last_collection, Links))
   if NotCount < WORKERS_COUNT:
       NotCount += 1
       with LinksLoaded:
           LinksLoaded.notify() 
   # else:
   #    time.sleep(2)
   #    NotCount = 0
             # LinksNotLoaded = False     

class TabInd:
            NAME = 0
            ARTICLE = 1
            BRAND = 2
            PRICE = 3
            SIZE = 4
            DESCRIPTION = 5
            PHOTO = 6
            LINK = 7  

#Парсинг страницы товара
def ParsingGoods(LinksLoaded, Locker):
     Goods = []
     Video = []
     # Вызов авторизации  
     print (f"ParsingGoods ({GoodsParser.name}): Ожидание")
     with LinksLoaded:
            LinksLoaded.wait()   
     print (f"ParsingGoods({GoodsParser.name}) Старт") 
     while not GoodsLinksQueue.empty():  
             CollectionName, Links = GoodsLinksQueue.get()
             print(f'ParsingGoods({GoodsParser.name}): Начат сбор данных товаров по коллекции {CollectionName}')  
             # Сбор данных со страницы товара в структуры и запись в список  
             # # Создание экземпляра драйвера             
             d=Service(ChromedriverPuth) 
             driver = webdriver.Chrome(service=d)
             # Загружаем cookie
             # with open('cookies.json', 'r') as f:
             #      cookies = json.load(f)
             #      s.cookies.update(cookies)
             # Вызов авторизации  
             # Authorization(driver) 
             for Link in Links:        
                driver.get(Link)  
                #Имя
                Name1 = driver.find_element(By.TAG_NAME,"h1").get_attribute("innerText")
                try:
                    Color = driver.find_element(By.XPATH,"//div[@class ='df jcsb rel']/span[2]").get_attribute("innerText")
                except:
                    Color = ""
                Name = f"{Name1} {Color}"
                #Артикл
                try:
                    Article = driver.find_element(By.XPATH,"//div[@class ='vendor df jcsb']/span[2]").get_attribute("innerText")
                except:
                    Article = ""
                #Цена
                Price = ""
                # driver.find_element(By.XPATH,"//span[@class='price']").get_attribute("innerText").strip("₽").replace("\xa0","")
                #Размер
                SizeList = []
                Sizes = driver.find_elements(By.XPATH,"//div[contains(@class,'offer-size size rel df aic jcc txt_bolder')][not(contains(@class,'disable'))]")
                if Sizes == []: continue
                else:
                    for Siz in Sizes: SizeList.append(Siz.get_attribute("data-size")) 
                    Size = ", ".join(SizeList)    
                # Описание 
                Description = "" 
                symbols = "АВ:"
                for sbl in symbols:
                     Composition = driver.find_element(By.XPATH,"//span[@class='compound txt_bolder']").get_attribute("innerText").replace(sbl, "") 
                try:
                     Discript = driver.find_element(By.XPATH,"//span[@class='description_text txt typing']").get_attribute("innerText")   
                except:
                    Discript = " Состав: "
                # -> Параметры модели и инструкции по уходу за изделием
                ModelSize, CareInstat = "", ""  
                ModUse = driver.find_elements(By.XPATH,'//div[@class="acc_item"]//div[@class="item_back txt"]')
                for ind, prms in enumerate(ModUse):  
                    if ind == 0: ModelSize = "Паметры модели: " + prms.get_attribute("innerText")[9:]
                    if ind == 1: CareInstract = "Правила ухода за изделием: " + prms.get_attribute("innerText")[9:]                      
                 # -> Таблица размеров (! Если нет размера: Наименование измерения  | XS  | S  | M  | L  | XL, Наименование измерения	84-92	88-96	92-100	96-104	100-108) 
                TabList = []
                ST = driver.find_elements(By.XPATH,"//table[@class='table']//tr")
                for st in ST:
                    DS = st.find_elements(By.XPATH, ".//td")
                    RD = [TS.get_attribute("innerText") for TS in DS] 
                    TabList.append("  | ".join(RD))
                TabSize = "\n".join(TabList)

                Description = Discript + " " + Composition + "\n" + ModelSize + "\n" + CareInstract + "; \n" + TabSize
                #Картинки
                Picture = []
                Pictures = driver.find_elements(By.XPATH,"//img[@alt='"+Name1+"']")
                if len(Pictures)==0:
                    Pictures = driver.find_elements(By.XPATH,"//img[@alt=' "+Name1+"']")
                if len(Pictures)==0:
                    print(Link)
                    continue
                # Video.append(Pictures[0].get_attribute("src"))
                for i in range(min(4, len(Pictures))):
                    response = requests.get(Pictures[i].get_attribute("src"))
                    response.raise_for_status()  # Проверяем успешность запроса
                    image = Image.open(BytesIO(response.content))
                    image = image.convert('RGB')
                    fname = Name1 + "img"+ str(i) +".jpg"
                    image.save("G:\\SP\\Hassfashion\\FotoHass\\" + fname, 'JPEG', quality=95)
                    Picture.append("https://hassfashion.neocities.org/hassfashion_test/" + fname.replace(" ", "%20"))  
                # Запись данных в экземляр структуры StructureOfProducts    
                StructureOfProduct = {
                    TabInd.NAME : Name,
                    TabInd.ARTICLE : Article,
                    TabInd.BRAND : 'Hass',
                    TabInd.PRICE : Price, 
                    TabInd.SIZE :Size,
                    TabInd.DESCRIPTION : Description,
                    TabInd.PHOTO : Picture,
                    TabInd.LINK : Link,
                    }
                # Запись структуры в список   
                Goods.append(StructureOfProduct)       
             # Передаем Goods 
             print(f'ParsingGoods ({GoodsParser.name}): Сбор данных товаров по коллекции {CollectionName} завершен')
             driver.quit()
             with Locker:
                RecordingToExcel(Goods, CollectionName)
             # print(f'ParsingGoods({GoodsParser.name}): Данные товаров по коллекции {CollectionName} записаны')
             Goods.clear()
             
#Запись в эксель      
def RecordingToExcel(Goods,CollectionName): 
    # Создание\загрузка эксель файла
    if os.path.exists(file_name):
        wb = load_workbook(file_path)
        print(f"RecordingInExcel: Файл '{file_name}' успешно загружен.")
        # Обращение к книге и листам  
        if CollectionName in wb.sheetnames:
            #Обращение к листу и удаление не актуальных данных  
            ws = wb[CollectionName] 
            print(f"RecordingInExcel: Лист '{CollectionName}' уже существует. Данные будут обновлены.")     
            if ws.cell(row=2, column=1).value is not None:
                for row in ws.iter_rows():
                    for cell in row: 
                        cell.value = None    
        else:
            #Создание листа 
            ws = wb.create_sheet(title=CollectionName)
            print(f"Создан новый лист '{CollectionName}'.")
    else:
       wb = Workbook()  
       print(f"RecordingInExcel:Файл '{file_name}' успешно создан.")     
       ws = wb.create_sheet(title=CollectionName)
       print(f"RecordingInExcel:Создан новый лист '{CollectionName}'.")
    # Добавление данных
    headers = ['Название', 'Артикл', 'Бренд', 'Цена', 'Размер', 'Описание', 'Изображение', 'Изображение1', 'Изображение2', 'Изображение3', 'Ссылка на товар']  
    for col_num, header in enumerate(headers, start=1):
         ws.cell(row=1, column=col_num, value=header)  
    for index, item in enumerate(Goods, start = 2 ):  
        ws.cell(row=index, column=1, value=item[TabInd.NAME])
        ws.cell(row=index, column=2, value=item[TabInd.ARTICLE])
        ws.cell(row=index, column=3, value=item[TabInd.BRAND])
        ws.cell(row=index, column=4, value=item[TabInd.PRICE])
        ws.cell(row=index, column=5, value=item[TabInd.SIZE])
        ws.cell(row=index, column=6, value=item[TabInd.DESCRIPTION])
        for imgindex, img in zip(range(4), item[TabInd.PHOTO]):   
             ws.cell(row=index, column=10 - imgindex, value = img) 
        ws.cell(row=index, column=11, value=item[TabInd.LINK ])
    wb.save(file_path)
    print(f"RecordingInExcel: Данные по категории {CollectionName} успешно записаны в файл '{file_name}'.")   
    try:  
        wb["Sheet"].title = "Лист1"
        wb.remove(wb["Лист1"])     
        wb.save(file_path)   
    except: pass 
      
if __name__ == '__main__': #Главная процедура 
    WORKERS_COUNT = 5
    GoodsLinksQueue = queue.Queue()
    LinksLoaded = threading.Condition()
    Locker = threading.Lock() 
    # s = requests.Session() # Создание сессии 
    Authorization() # Вызов авторизации 
    # Вызов производящего потока
    linksParser = threading.Thread(target = GettingLinks, args = (LinksLoaded,))
    linksParser.start() 
    # Вызов группы потребляющих потоков
    for i in range(WORKERS_COUNT):
        GoodsParser = threading.Thread(name = str(i), target = ParsingGoods, args = (LinksLoaded, Locker))
        GoodsParser.start()
    for thread in threading.enumerate():
            if thread != threading.main_thread():  
                thread.join()  
    linksParser.join()
    # driver.quit()
    print("Готово") 
    input()
    
    # GoodsParser = threading.Thread(target = ParsingGoods, args = (LinksLoaded,driver))
    # GoodsParser.start()
    # GoodsParser.join()
