from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import threading
import queue
import time

import openpyxl
from openpyxl import Workbook,load_workbook
import os

# Указание пути компонентам парсера
# Определение пути к chromedriver 
ChromedriverPuth = 'G:\\NRU\\SP\\Parsers\\selenium\\chromedriver\\win64\\136.0.7103.92\\chromedriver.exe'
# Определение пути к файлу с коллекциями HassCatalog   
CollectionFilePath = ".\HassCatalog.xlsx" 
# Определение имени и автоформирование относительного пути к файлу загрузки данных HassDate
file_name = "HassDate.xlsx"
file_path = "./" + file_name  

# Авторизация на ХассФэсшен 
def Authorization(driver):
    driver.get("https://hassfashion.ru/auth/?login=yes")    
    Auth_Log = driver.find_element(By.NAME, "USER_LOGIN")
    Auth_Log.send_keys('Olga_Guschina@mail.ru')
    Auth_Password = driver.find_element(By.NAME, "USER_PASSWORD")
    Auth_Password.send_keys('optopt')
    Auth_Imp = driver.find_element(By.NAME, "Login")
    Auth_Imp.click()  
    time.sleep(10) 
    
#Получение страниц каталога из Экселя
def GettingLinks(LinksLoaded):
    #Создание экземпляра драйвера
    s=Service(ChromedriverPuth) 
    driver = webdriver.Chrome(service=s)
    # Обращение к эксель файлу 
    wb = load_workbook(CollectionFilePath)
    ws = wb.active 
    LinkPages = [] 
    last_collection = ws[2][0].value  
    for row in range(2, ws.max_row+1):  
        current_collection = ws[row][0].value 
        if current_collection == None: break  
        if last_collection !=  current_collection : 
             ParsingCollection(driver, LinksLoaded, LinkPages, last_collection)             
        LinkPages.append(ws[row][2].value) 
        last_collection = current_collection
    ParsingCollection(driver, LinksLoaded, LinkPages, last_collection)
    driver.quit() 

def ParsingCollection(driver, LinksLoaded, LinkPages, last_collection):
   print(f'GettingLinks: Начат сбор ссылок для коллекции {last_collection}')   
   Links = [] 
   FilteredPaginatorConteyner = []
   for LinkPage in LinkPages: 
        driver.get(LinkPage)  
        PaginatorPage = LinkPage 
        while True:
            print(f"LinkPage: {PaginatorPage}")
            # Отбор товаров в наличие                        
            OnSales = driver.find_elements(By.XPATH, "//div[@class='res_cards']/div[not(@style)]") 
            for OnSale in OnSales:   
                a = OnSale.find_element(By.TAG_NAME, "a")
                Links.append(a.get_attribute("href"))     
            try:
                PaginatorPage = driver.find_element(By.XPATH, "//div[contains(@data-pagination,'PAGEN')]")   
                PaginatorLoadPage = "https://hassfashion.ru" + driver.find_element(By.XPATH, "//div[contains(@data-pagination,'PAGEN')]").get_attribute("data-href") 
                driver.get(PaginatorLoadPage)           
            except: break        
   LinkPages.clear()        
   print(f'GettingLinks: Закончен сбор ссылок для коллекции {last_collection}')  
   GoodsLinksQueue.put((last_collection, Links))
   with LinksLoaded:
            LinksLoaded.notify() 
            LinksNotLoaded = False     
   
class TabInd:
            NAME = 0
            ARTICLE = 1
            BRAND = 2
            PRICE = 3
            SIZE = 4
            DESCRIPTION = 5
            PHOTO = 6
            LINK = 7  

#Парсинг страницы товара
def ParsingGoods(LinksLoaded, driver):
  Goods = []
  Video = []
  print ("ParsingGoods Ожидание")
  with LinksLoaded:
        LinksLoaded.wait()
  print ("ParsingGoods Старт")   
  while not GoodsLinksQueue.empty():  
     CollectionName, Links = GoodsLinksQueue.get()
     print(f'ParsingGoods: Начат сбор данных товаров по коллекции {CollectionName}')  
     # Сбор данных со страницы товара в структуры и запись в список 
     for Link in Links:
        driver.get(Link)         
        #Имя
        Name1 = driver.find_element(By.XPATH,"//div[@class ='title h3 mobile']").get_attribute("innerText")
        Color = driver.find_element(By.XPATH,"//div[@class ='df jcsb rel']/span[2]").get_attribute("innerText")
        Name = f"{Name1} {Color}"
        #Артикл
        Article = driver.find_element(By.XPATH,"//div[@class ='vendor df jcsb']/span[2]").get_attribute("innerText")
        #Цена
        Price = driver.find_element(By.XPATH,"//span[@class='price']").get_attribute("innerText").strip("₽").replace("\xa0","")
        #Размер
        SizeList = []
        Sizes = driver.find_elements(By.XPATH,"//div[contains(@class,'offer-size size df aic jcc txt_bolder')][not(contains(@class,'disable'))]")
        if Sizes == []: Size = "-"
        else:
            for Siz in Sizes: SizeList.append(Siz.get_attribute("innerText")) 
            Size = ", ".join(SizeList) + "."    
        # Описание 
        Description = "" 
        symbols = "АВ:"
        for sbl in symbols:
             Composition = driver.find_element(By.XPATH,"//span[@class='compound txt_bolder']").get_attribute("innerText").replace(sbl, "") 
        try:
             Discript = driver.find_element(By.XPATH,"//span[@class='description_text txt typing']").get_attribute("innerText")   
        except:
            Discript = " Состав: "
        Description = Discript + " " + Composition 
        #Картинки
        Picture = []
        Pictures = driver.find_elements(By.XPATH,"//div[@class ='miniatures_item df fdc']//img")
        # Video.append(Pictures[0].get_attribute("src"))
        if len(Pictures) > 5:
            for index, Pict in enumerate(Pictures):
                if index > 0:           
                    if index % 2 != 0:
                         Picture.append(Pict.get_attribute("src"))
        else: 
             for index, Pict in enumerate(Pictures):
                if index > 0:    
                    Picture.append(Pict.get_attribute("src"))    
        # Запись данных в экземляр структуры StructureOfProducts    
        StructureOfProduct = {
            TabInd.NAME : Name,
            TabInd.ARTICLE : Article,
            TabInd.BRAND : 'Hassfastion',
            TabInd.PRICE : Price, 
            TabInd.SIZE :Size,
            TabInd.DESCRIPTION : Description,
            TabInd.PHOTO : Picture,
            TabInd.LINK : Link,
            }
        # Запись структуры в список   
        Goods.append(StructureOfProduct)       
     # Передаем Goods 
     print(f'ParsingGoods: Сбор данных товаров по коллекции {CollectionName} завершен')
     RecordingToExcel(Goods, CollectionName)
     Goods.clear()
       
#Запись в эксель      
def RecordingToExcel(Goods,CollectionName): 
    # Создание\загрузка эксель файла
    if os.path.exists(file_name):
        wb = load_workbook(file_path)
        print(f"RecordingInExcel: Файл '{file_name}' успешно загружен.")
        # Обращение к книге и листам  
        if CollectionName in wb.sheetnames:
            #Обращение к листу и удаление не актуальных данных  
            ws = wb[CollectionName] 
            print(f"RecordingInExcel: Лист '{CollectionName}' уже существует. Данные будут обновлены.")     
            if ws.cell(row=2, column=1).value is not None:
                for row in ws.iter_rows():
                    for cell in row: 
                        cell.value = None    
        else:
            #Создание листа 
            ws = wb.create_sheet(title=CollectionName)
            print(f"Создан новый лист '{CollectionName}'.")
    else:
       wb = Workbook()  
       print(f"RecordingInExcel:Файл '{file_name}' успешно создан.")     
       ws = wb.create_sheet(title=CollectionName)
       print(f"RecordingInExcel:Создан новый лист '{CollectionName}'.")
    # Добавление данных
    headers = ['Название', 'Артикл', 'Бренд', 'Цена', 'Размер', 'Описание', 'Изображение', 'Изображение1', 'Изображение2', 'Изображение3', 'Ссылка на товар']  
    for col_num, header in enumerate(headers, start=1):
         ws.cell(row=1, column=col_num, value=header)  
    for index, item in enumerate(Goods, start = 2 ):  
        ws.cell(row=index, column=1, value=item[TabInd.NAME])
        ws.cell(row=index, column=2, value=item[TabInd.ARTICLE])
        ws.cell(row=index, column=3, value=item[TabInd.BRAND])
        ws.cell(row=index, column=4, value=item[TabInd.PRICE])
        ws.cell(row=index, column=5, value=item[TabInd.SIZE])
        ws.cell(row=index, column=6, value=item[TabInd.DESCRIPTION])
        for imgindex, img in enumerate(reversed(item[TabInd.PHOTO])):   
             if imgindex < 4:
                ws.cell(row=index, column=7 + imgindex, value = img) 
        ws.cell(row=index, column=11, value=item[TabInd.NAME])
    wb.save(file_path)
    print(f"RecordingInExcel: Данные по категории {CollectionName} успешно записаны в файл '{file_name}'.")   
    try:  
        wb["Sheet"].title = "Лист1"
        wb.remove(wb["Лист1"])     
        wb.save(file_path)   
    except: pass 
      
#Главная процедура     
GoodsLinksQueue = queue.Queue()
LinksLoaded = threading.Condition()
##Создание экземпляра драйвера  
d=Service(ChromedriverPuth)  
driver = webdriver.Chrome(service=d)
#Вызов авторизации  
Authorization(driver) 
linksParser = threading.Thread(target = GettingLinks, args = (LinksLoaded,))
linksParser.start()
GoodsParser = threading.Thread(target = ParsingGoods, args = (LinksLoaded,driver))
GoodsParser.start()
linksParser.join() 
GoodsParser.join()
driver.quit()
print("Готово") 
input()

