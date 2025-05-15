
  
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import threading
import queue
import time

import openpyxl
from openpyxl import Workbook,load_workbook
import xlsxwriter
import os


 # Авторизация на ХассФэсшен 
def AuthorizationForHasfesshen(driver):
    driver.get("https://hassfashion.ru/auth/?login=yes")
    time.sleep(5)

    Auht_Log = driver.find_element(By.NAME, "USER_LOGIN")
    Auht_Log.send_keys('Olga_Guschina@mail.ru')

    Auht_Password = driver.find_element(By.NAME, "USER_PASSWORD")
    Auht_Password.send_keys('optopt')

    Auth_Imp = driver.find_element(By.NAME, "Login")
    Auth_Imp.click()
    
    time.sleep(2) 
          
#Получение страниц каталога из Экселя
def GettingСategories(driver):
    #Создание экземпляра драйвера
    # s=Service('G:\\NRU\\SP\\Parsing\\selenium\\chromedriver\\win64\\136.0.7103.92\\chromedriver.exe')
    # driver = webdriver.Chrome(service=s)
    
    # Обращение к эксель файлу
    file_path = "G:\\NRU\\SP\\Parsing\\TestSelenium\\TestSelenium\\HassCatalog.xlsx"  
    wb = load_workbook(file_path)
    ws = wb.active 
    LinkPages = []
    for index, row in enumerate(range(2, ws.max_row + 1)):
        current_collection = ws[row][0].value 
        if index > 0:
            last_collection = ws[row - 1][0].value  
            if last_collection !=  current_collection and last_collection !=  None :
                    print(f'Начат сбор данных для следующей коллекции: {last_collection}')
                    Goods = [] 
                    Video = []
                    for LinkPage in LinkPages:   
                         driver.get(LinkPage) 
                         print(LinkPage)
                         time.sleep(5)  
                    # Отбор товаров в наличие
                    Links = [] 
                    OnSales = driver.find_elements(By.XPATH, "//div[@class='res_cards']/div[not(@style)]") 
                    for OnSale in OnSales:   
                           a = OnSale.find_element(By.TAG_NAME, "a")
                           Links.append(a.get_attribute("href"))   
                    
                    time.sleep(10)
                    # LinksLoaded.notify()
                    ParsingPage(Links,last_collection)                
                    LinkPages = [] 
                    
        LinkPages.append(ws[row][2].value) 
        
    lastest_collection = ws[ws.max_row][0].value
    print(f'Начат сбор данных для следующей коллекции: {lastest_collection}')
    for LinkPage in LinkPages:   
           driver.get(LinkPage) 
           print(LinkPage)
           time.sleep(5)  
    # Отбор товаров в наличие
    Links = [] 
    OnSales = driver.find_elements(By.XPATH, "//div[@class='res_cards']/div[not(@style)]") 
    for OnSale in OnSales:   
         a = OnSale.find_element(By.TAG_NAME, "a")
         Links.append(a.get_attribute("href"))   
                    
    time.sleep(10) 
    # LinksLoaded.notify()
    ParsingPage(Links,lastest_collection) 
    
class TabInd:
            NAME = 0
            ARTICLE = 1
            BRAND = 2
            PRICE = 3
            SIZE = 4
            DESCRIPTION = 5
            PHOTO = 6
            LINK = 7  
#Парсинг страницы
def ParsingPage(Links,CollectionName):

   Goods = []
   Video = []
  #Создание экземпляра драйвера
    # d=Service('G:\\NRU\\SP\\Parsing\\selenium\\chromedriver\\win64\\136.0.7103.92\\chromedriver.exe')
    # driver = webdriver.Chrome(service=d)

    # Collectioname = CollectionName
    #Вызов авторизации  
   AuthorizationForHasfesshen(driver)   
# Сбор данных со страницы товара в структуры и запись в список 
   for Link in Links:
        driver.get(Link)         
        #Имя
        Name1 = driver.find_element(By.XPATH,"//div[@class ='title h3 mobile']").get_attribute("innerText")
        Color = driver.find_element(By.XPATH,"//div[@class ='df jcsb rel']/span[2]").get_attribute("innerText")
        Name = f"{Name1} {Color}"
        #Артикл
        Article = driver.find_element(By.XPATH,"//div[@class ='vendor df jcsb']/span[2]").get_attribute("innerText")
        #Цена
        Price = driver.find_element(By.XPATH,"//span[@class='price']").get_attribute("innerText").strip("₽").replace("\xa0","")
        #Размер
        SizeList = []
        Sizes = driver.find_elements(By.XPATH,"//div[contains(@class,'offer-size size df aic jcc txt_bolder')][not(contains(@class,'disable'))]")
        if Sizes == []: Size = "-"
        else:
            for Siz in Sizes: SizeList.append(Siz.get_attribute("innerText")) 
            Size = ", ".join(SizeList) + "."    
        # Описание 
        Description = "" 
        symbols = "АВ:"
        for sbl in symbols:
             Composition = driver.find_element(By.XPATH,"//span[@class='compound txt_bolder']").get_attribute("innerText").replace(sbl, "") 
        try:
             Discript = driver.find_element(By.XPATH,"//span[@class='description_text txt typing']").get_attribute("innerText")   
        except:
            Discript = " Состав: "
        Description = Discript + " " + Composition 
        #Картинки
        Picture = []
        Pictures = driver.find_elements(By.XPATH,"//div[@class ='miniatures_item df fdc']//img")
        Video.append(Pictures[0].get_attribute("src"))
        for index, Pict in enumerate(Pictures):
            if len(Pictures) > 5 and index > 0:           
                    if index % 2 != 0:
                         Picture.append(Pict.get_attribute("src"))
            elif index > 0:                  
                    Picture.append(Pict.get_attribute("src"))    
        # Запись данных в экземляр структуры StructureOfProducts    
        StructureOfProduct = {
            TabInd.NAME : Name,
            TabInd.ARTICLE : Article,
            TabInd.BRAND : 'Hassfastion',
            TabInd.PRICE : Price, 
            TabInd.SIZE :Size,
            TabInd.DESCRIPTION : Description,
            TabInd.PHOTO : Picture,
            TabInd.LINK : Link,
            }
        # Запись структуры в список   
        Goods.append(StructureOfProduct)
  # Передаем Goods
   RecordingInExcel(Goods, CollectionName)
    
#Запись в эксель      
def RecordingInExcel(Goods,CollectionName): 
    # Создание\загрузка эксель файла
    file_name = "hasf-parser2.xlsx"
    file_path = "G:\\NRU\\SP\\Parsing\\TestSelenium\\TestSelenium\\hasf-parser2.xlsx" 
    if os.path.exists(file_name):
        wb = load_workbook(file_path)
        print(f"Файл '{file_name}' успешно загружен.")
        # Обращение к книге и листам  
        if CollectionName in wb.sheetnames:
            #Обращение к листу и удаление не актуальных данных  
            ws = wb[CollectionName] 
            print(f"Лист '{CollectionName}' уже существует. Данные будут обновлены.")
            if ws.cell(row=2, column=1).value is not None:
                for row in ws.iter_rows():
                    for cell in row: 
                        cell.value = None    
        else:
            #Создание листа 
            ws = wb.create_sheet(title=CollectionName)
            print(f"Создан новый лист '{CollectionName}'.")
    else:
       wb = Workbook()  
       print(f"Файл '{file_name}' успешно создан.")     
       ws = wb.create_sheet(title=CollectionName)
       print(f"Создан новый лист '{CollectionName}'.")
   
    
    # Добавление данных
    headers = ['Название', 'Артикл', 'Бренд', 'Цена', 'Размер', 'Описание', 'Изображение', 'Изображение1', 'Изображение2', 'Изображение3', 'Ссылка на товар']  
    for col_num, header in enumerate(headers, start=1):
         ws.cell(row=1, column=col_num, value=header)
    
    for index, item in enumerate(Goods, start=2):  
        ws.cell(row=index, column=1, value=item[0])  
        ws.cell(row=index, column=2, value=item[1])
        ws.cell(row=index, column=3, value=item[2])
        ws.cell(row=index, column=4, value=item[3])
        ws.cell(row=index, column=5, value=item[4])
        ws.cell(row=index, column=6, value=item[5])
        for imgindex, img in enumerate(reversed(item[6])):   
             if imgindex < 4:
                ws.cell(row=index, column=7 + imgindex, value = img) 
        ws.cell(row=index, column=11, value=item[7])
    wb.save(file_path)
    print(f"Данные по категории {CollectionName} успешно записаны в файл '{file_name}'.") 
    
    try:  
        wb["Sheet"].title = "Лист1"
        wb.remove(wb["Лист1"])     
        wb.save(file_path)   
    except: pass 
    

    
#Главная процедура  
s=Service('G:\\NRU\\SP\\Parsing\\selenium\\chromedriver\\win64\\136.0.7103.92\\chromedriver.exe')
driver = webdriver.Chrome(service=s) 
GettingСategories(driver)
# GoodsLinksQueue = queue.Queue()

# LinksLoaded = threading.Condition()
# links_parser = threading.Thread(target = GettingСategories(), args = (LinksLoaded,))
# links_parser.start()



# book_locker = threading.Lock()

# for i in range(WORKERS_COUNT):
#      goods_parser = threading.Thread(name = str(i), target = parse_goods, args = (links_loaded, book_locker))
#      goods_parser.start()

# for thread in threading.enumerate():
#      if thread != threading.main_thread():
#          thread.join()

# g_book.close()
# print("Готово")
# a = input()
    
print("Готово")
input()



    
